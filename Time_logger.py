from os import path
from datetime import datetime
from keyboard import wait
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from win10toast import ToastNotifier

toaster = ToastNotifier()
filename = 'work_log.xlsx'
delta_time_zero = datetime.now() - datetime.now()
current_ts_row  = 2
last_day_row = 1
last_day = ""
last_day_value = delta_time_zero
working = False
columns_names = {'Date':'A', 'Working Time':'B', 'project name':'C', 'project time':'D', 'Start':'E', 'End':'F', 'Difference':'G'}

def write_aligned(idx, txt):
	line_num = int(idx[1:])
	lines_to_add = int(line_num / 1000) *2 # there is no A0 and A1 is reserved
	idx = idx[0] + str(line_num %1000 + lines_to_add)
	ws[idx] = txt
	ws[idx].alignment = Alignment(horizontal='center')

def add_new_entry():
	
	time_worked = ""
	
	global current_ts_row
	global last_day_row
	global last_day
	global last_day_value
	global working
	global start_time
	global columns_names
	full_date = '%Y-%m-%d %H:%M:%S'

	# Get the current timestamp
	timestamp_now = datetime.now()
	today = str(timestamp_now.strftime('%Y-%m-%d'))
	
	# Write the timestamp and key combination to the worksheet 
	if not working:
		write_aligned(columns_names['Start']+str(current_ts_row) , timestamp_now.strftime(full_date))
		working = True
		start_time = timestamp_now
		
	else:

		working = False		
		start_day = str(start_time.strftime('%Y-%m-%d'))
		
		if start_day != today:
			new_end_time = datetime.strptime(start_day+" 23:59:59",full_date)

			write_aligned(columns_names['End']+str(current_ts_row) , new_end_time.strftime(full_date))
			diff = new_end_time - start_time + datetime.strptime("1","%S")
			write_aligned(columns_names['Difference']+str(current_ts_row) , str(diff).split(" ")[1])
			current_ts_row += 1

			if start_day == last_day:
				last_day_value += diff
				write_aligned(columns_names['Working Time']+str(last_day_row), str(last_day_value).split(" ")[1])

			else:
				last_day_row += 1
				write_aligned(columns_names['Date']+str(last_day_row), start_day)
				write_aligned(columns_names['Working Time']+str(last_day_row), str(diff).split(" ")[1])
				last_day = start_day
				last_day_value = diff


			new_start_time = datetime.strptime(today+" 0:0:0",full_date)
			write_aligned(columns_names['Start']+str(current_ts_row) , new_start_time)
			write_aligned(columns_names['End']+str(current_ts_row) , timestamp_now.strftime(full_date))
			diff = timestamp_now - new_start_time
			write_aligned(columns_names['Difference']+str(current_ts_row) , diff)
			current_ts_row += 1

			last_day_row += 1
			write_aligned(columns_names['Date']+str(last_day_row), today)
			write_aligned(columns_names['Working Time']+str(last_day_row), diff)
			last_day = today
			last_day_value = diff

		else:

			write_aligned(columns_names['End']+str(current_ts_row) , timestamp_now.strftime(full_date))
			diff = timestamp_now - start_time
			write_aligned(columns_names['Difference']+str(current_ts_row) , diff)
			current_ts_row += 1

			if today == last_day:
				last_day_value += diff
				write_aligned(columns_names['Working Time']+str(last_day_row), last_day_value)

			else:
				last_day_row += 1
				write_aligned(columns_names['Date']+str(last_day_row), today)
				write_aligned(columns_names['Working Time']+str(last_day_row), diff)
				last_day = today
				last_day_value = diff
			

			time_worked = str(last_day_value).split(".")[0] 
			print("time worked today = " + time_worked)
   
			# just a procastination counter measure i hope it works ;)
			print(last_day_value, type(last_day_value))
			if int(str(last_day_value).split(':')[0]) > 8 :
				print("you may go to sleep now ;)")
			else:
				try:
					print("time to finish = " + str(datetime.strptime("8","%H") - last_day_value).split(".")[0].split(" ")[1]) 
				except:
					pass

	# Save the changes to the workbook 
	try:
		wb.save(filename)
	except:
		toaster.show_toast("ERROR please close the excel file then try again", " ", duration=1, threaded=True)
		print("close the file")
		exit(1)
	
	return time_worked

def prepare_file():

	global wb
	global ws
	global current_ts_row
	global last_day_row
	global last_day
	global last_day_value
	# Check if the file exists
	if path.isfile(filename):
			
		# load the workbook and select the active worksheet
		wb = load_workbook(filename)
		ws = wb.active

		for i in range(2,1001):

			# wait until fist empty cell 
			try:
				str(ws[columns_names['Start']+str(i)].value[0])
			except:
				current_ts_row = i
				break

		for i in range(2,1001):

			if ws[columns_names['Date']+str(i)].value == None:
				last_day_row = i -1
				if i != 2:
					last_day = ws[columns_names['Date']+ str(i-1)].value
					last_day_value = ws[columns_names['Working Time']+ str(i-1)].value
				break

	else:

		# If the file doesn't exist, create a new workbook and worksheet
		wb = Workbook()
		ws = wb.active

		# Add headers to the worksheet 
		for key, value in columns_names.items():
			write_aligned(value+'1',key)
			ws.column_dimensions[value].width = 23

def check_the_file():
	try:
		wb.save(filename)
		toaster.show_toast("Work logger is Ready", " ", duration=1, threaded=True)
		print("ready")
	except:
		toaster.show_toast("ERROR please close the excel file then try again", " ", duration=1, threaded=True)
		print("close the file then try again")
		exit(1)

def listen_and_run():
	# Start the keylogger 
	while True:
		try:
			# Wait for a key combination to be pressed 
			wait(r'q+1')
		
			time_worked = add_new_entry()
			
			if working:
				text1 = "START WORKING"
				text2 = "started"
			else:
				text1 = "STOP WORKING    worked today: "+time_worked
				text2 = "stopped"
				
			toaster.show_toast(text1, " ", duration=1, threaded=True)
			print(text2)
				
		except KeyboardInterrupt:
			# If the user presses Ctrl+C, exit the loop 
			if working:	add_new_entry()
			exit()

prepare_file()
check_the_file()
listen_and_run()
